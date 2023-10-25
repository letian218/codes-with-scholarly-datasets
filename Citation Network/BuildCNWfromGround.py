# -*- coding: utf-8 -*-
"""
Created on Sun May  9 11:37:56 2021
Code for processing a WoS dataset contains many Full Record information files (txt format)
Some user-defined functions were not used but have been preserved
@author: Letian
"""

import os
import openpyxl as op
import pandas as pd
import multiprocessing
import time


def mp_lize(func, listargs):
    multiprocessing.freeze_support()
    cores = multiprocessing.cpu_count()
    print('\n'+time.asctime(time.localtime(time.time())), end=': ')
    print('pool Starting...\n')
    pool = multiprocessing.Pool(processes=cores)
    results = []
    for args in listargs:
        results.append(pool.apply_async(func, args=args))
    pool.close()
    pool.join()
    print('\n'+time.asctime(time.localtime(time.time())), end=': ')
    print('pool closed and joined...\n')
    items = []
    for item in results:
        items.append(item.get())
    del item
    del results
    del pool
    return items


def __user_warning(tip='We need to talk...'):
    print('Warning: {}'.format(tip))
    if input('Ignore it and continue? [Y/N]:') == ('y' or 'Y'):
        pass
    else:
        raise UserWarning(tip)


def iter_upper(str_iterable):
    return [str(item).upper() for item in str_iterable]


def file_getter(path):
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            yield os.path.join(root, file)


def file_getter_mpv(path):
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            yield (os.path.join(root, file),)


def txt_spliter(file, encoding='utf-8', start_tag='PT', end_tag='ER'):
    f = open(file, encoding=encoding)
    file_txt = f.readlines()
    f.close()
    raw_txts = []
    start = 0
    end = 0
    for i in range(len(file_txt)):
        if file_txt[i][:2] == start_tag:
            start = i
        elif file_txt[i][:2] == end_tag:
            end = i
        if start*end:
            raw_txts.append(file_txt[start:end])
            start = 0
            end = 0
    if start+end:
        return None
    elif raw_txts == []:
        return None
    return raw_txts


def paper_extractor(raw_txt):
    paper_inform = {}
    for line in raw_txt:
        if line[:2] != '  ':
            tag = line[:2]
        if tag in paper_inform:
            paper_inform[tag].append(line[3:-1])
        else:
            paper_inform[tag] = [line[3:-1]]
    return paper_inform


def file_reader(file, encoding='utf-8'):
    fsp = txt_spliter(file, encoding=encoding)
    if fsp is None:
        print(file)
        raise UserWarning('TagError. Check file or tags.')
    orders = os.path.basename(file)[:-4].split('-')
    if len(fsp) != (int(orders[1])-int(orders[0])+1):
        print(orders)
        print(file)
        raise UserWarning('{} OrderError. Check file or filename.'.format(len(fsp)))
    file_inform = []
    for paper in fsp:
        file_inform.append(paper_extractor(paper))
    return file_inform


def path_reader(src_path, encoding='utf-8'):
    path_inform = []
    uts = []
    for file in file_getter(src_path):
        for paper_inform in file_reader(file, encoding=encoding):
            path_inform.append(paper_inform)
            uts.append(paper_inform['UT'][0][4:])
        if len(uts) != len(set(uts)):
            print(file, len(uts), len(set(uts)))
            raise UserWarning('UTsError. Check file.')
    return path_inform


def path_reader_2(src_path, encoding='utf-8'):
    for file in file_getter(src_path):
        for paper_inform in file_reader(file, encoding=encoding):
            yield paper_inform


def ut_checker(src_path):
    uts = {}
    for file in file_getter(src_path):
        for paper_inform in file_reader(file):
            ut = paper_inform['UT'][0][4:]
            if ut in uts:
                print(ut)
                print(file)
                print(uts[ut])
            else:
                uts[ut] = file


def data_exporter(src_path, out_path):
    if not os.path.exists(out_path):
        os.makedirs(out_path)
    # uts = []
    wb = op.workbook.Workbook()
    ws = wb.active
    order = 0
    n = 0
    tags = []
    for file in file_getter(src_path):
        for paper_inform in file_reader(file):
            if n >= 100000:
                wb.save(os.path.join(out_path, 'Data_Extraction_{}.xlsx'.format(order)))
                order += 1
                n = 0
                wb = op.workbook.Workbook()
                ws = wb.active
                for i in range(len(tags)):
                    ws.cell(row=1, column=i+1).value = tags[i]
                del i
            # uts.append(paper_inform['UT'][0][4:])
            for tag in paper_inform:
                if tag in tags:
                    inform = ' || '.join(paper_inform[tag])
                    ws.cell(row=n+2, column=tags.index(tag)+1).value = inform
                else:
                    tags.append(tag)
                    ws.cell(row=1, column=len(tags)).value = tag
                    inform = ' || '.join(paper_inform[tag])
                    ws.cell(row=n+2, column=len(tags)).value = inform
            n += 1
        '''
        if len(uts) != len(set(uts)):
            print(file)
            raise UserWarning('UTsError. Check file.')
        '''
    wb.save(os.path.join(out_path, 'Data_Extraction_{}.xlsx'.format(order)))


def essence_exporter(src_path, out_path, essence_key, file_name='Extraction', sep=', ', re_k='NA', re_d={}):
    if not os.path.exists(out_path):
        os.makedirs(out_path)
    f = open(os.path.join(os.path.join(out_path, '{}.csv'.format(file_name))),
             mode='w',
             encoding='utf-8')
    line = ''
    for key in essence_key:
        line += '{}{}'.format(key, sep)
    f.write(line[:-len(sep)]+'\n')
    for file in file_getter(src_path):
        for paper_inform in file_reader(file):
            ut = paper_inform['UT'][0][4:]
            paper_inform['UT'][0] = ut
            if ut in re_d:
                paper_inform[re_k] = [re_d[ut]]
            line = ''
            for key in essence_key:
                if key in paper_inform:
                    line += '{}{}'.format(' || '.join(paper_inform[key]), sep)
                else:
                    line += sep
            f.write(line[:-len(sep)]+'\n')
    f.close()


def batch_writer(target_path, order, size, content):
    if not os.path.exists(target_path):
        os.makedirs(target_path)
    txt_file_name = '{}-{}.txt'.format(1+order*size, (1+order)*size)
    f = open(os.path.join(target_path, txt_file_name),
             mode='w',
             encoding='utf-8')
    f.write(content)
    f.close()


def batch_writer_lite(target_path, name, content):
    if not os.path.exists(target_path):
        os.makedirs(target_path)
    txt_file_name = '{}.txt'.format(name)
    f = open(os.path.join(target_path, txt_file_name),
             mode='w',
             encoding='utf-8')
    f.write(content)
    f.close()


def max_usrdef(dict_like, exception=False, key_exception=False):
    keys = [key for key in dict_like]
    if exception:
        dump = []
        for key in keys:
            if dict_like[key] in exception:
                dump.append(key)
        for item in dump:
            keys.remove(item)
    if key_exception:
        for item in key_exception:
            keys.remove(item)
    max_item = [[keys[0]], dict_like[keys[0]]]
    for i in range(1, len(keys)):
        if dict_like[keys[i]] < max_item[1]:
            continue
        elif dict_like[keys[i]] > max_item[1]:
            max_item = [[keys[i]], dict_like[keys[i]]]
        else:
            max_item[0].append(keys[i])
    return max_item


def pop_max_item(iterable):
    m = list(iterable.values())[0]
    for key in iterable:
        if iterable[key] >= m:
            m = iterable[key]
            target = key
    return (target, iterable.pop(target))


def cr_refinery(cr_txt, jn_sub={}):
    '''
    cr_inform = [Authors, ESI_Sub, Pub_Name, Pub_Year, DOI]
    '''
    items = cr_txt.split(', ')
    cr_inform = ['', '', '', '', '']
    checked_items = []
    for item in items:
        if 'DOI ' in item:
            checked_items.append(item)
            doi = item[4:]
            while 'DOI ' in doi:
                doi = doi[4:]
            cr_inform[-1] = doi
        elif item in jn_sub:
            cr_inform[1] = jn_sub[item][0]
            cr_inform[2] = jn_sub[item][1]
            checked_items.append(item)
        elif len(item) == 4:
            try:
                year = int(item)
            except ValueError:
                continue
            if year in range(1900, 2019):
                cr_inform[-2] = str(year)
                checked_items.append(item)
    del item
    if items[0] not in checked_items:
        cr_inform[0] = items[0]
    return tuple(cr_inform)


def inform_reader(file, *tags):
    wb = op.load_workbook(file)
    ws = wb.active
    inform = [''] * len(tags)
    for row in ws.rows:
        tag = row[0].value
        if tag in tags:
            inform[tags.index(tag)] = [item.value for item in row[1:] if item.value]
    return inform


def inform_reader_lite(ws, *tags):
    inform = [''] * len(tags)
    for row in ws.rows:
        tag = row[0].value
        if tag in tags:
            inform[tags.index(tag)] = [item.value for item in row[1:] if item.value]
    return inform


def keyword_reader(path):
    keys = {}
    titles = []
    for paper_inform in path_reader(path):
        if 'DE' in paper_inform:
            for kw in ' '.join(paper_inform['DE']).split('; '):
                kw = kw.upper()
                if kw in keys:
                    keys[kw] += 1
                else:
                    keys[kw] = 1
        elif 'TI' in paper_inform:
            titles.append(' '.join(paper_inform['TI']))
    all_title = ' || '.join(titles).upper()
    for kw in keys:
        if kw in all_title:
            keys[kw] += 1
    return keys


def xlsx_key(ws, key, key_type='col'):
    if key_type == 'col':
        for i in range(ws.max_column):
            if ws.cell(row=1, column=i+1).value == key:
                return i+1
    elif key_type == 'row':
        for i in range(ws.max_row):
            if ws.cell(row=i+1, column=1).value == key:
                return i+1
    else:
        print(key_type)
        raise UserWarning('Key_Type Error. Check it out.')


def paper_format(paper_inform):
    raw_txt = ''
    for tag in paper_inform:
        raw_txt += tag + ' ' + '\n   '.join(paper_inform[tag]) + '\n'
    return raw_txt + 'ER\n\n'


def buffer_writer(buffer, out_path, file_name, encode='utf-8'):
    if not os.path.exists(out_path):
        os.makedirs(out_path)
    f = open(os.path.join(out_path, file_name),
             mode='w',
             encoding=encode)
    f.write("FN Enhanced WoS' Data\nVR 1.1\n")
    for paper_inform in buffer:
        f.write(paper_format(paper_inform))
    f.write('EF\n')
    f.close()


def doi_match(paper_inform, cr):
    cr = cr.upper()
    if 'DOI ' in cr:
        if 'DI' in paper_inform:
            if paper_inform['DI'][0].upper() in cr:
                return True  # DOI matched.
            else:
                return False  # Wrong DOI.
        else:
            return None  # DOI missing in original data.
    else:
        return None


def au_match(paper_inform, cr):
    cr = cr.upper()
    if 'AU' in paper_inform:
        fau = ' '.join(paper_inform['AU'][0].split(', ')) + ', ' 
        if fau.upper() in cr:
            return True  # Au matched.
        else:
            return False  # Wrong AU.
    else:
        return None


def vl_match(paper_inform, cr):
    cr = cr.upper()
    cr_vl = None
    for item in cr.split(', '): 
        if (item[0] == 'V') and (item[1:].isdigit()):
            cr_vl = item[1:]
    if cr_vl is None:
        return None
    elif 'VL' in paper_inform:
        vl = paper_inform['VL'][0]
        if vl == cr_vl:
            return True  # VL matched.
        else:
            return False  # Wrong VL.
    else:
        return None


def bp_match(paper_inform, cr):
    cr = cr.upper()
    cr_bp = None
    for item in cr.split(', '):
        if (item[0] == 'P') and (item[1:].isdigit()):
            cr_bp = item[1:]
    if cr_bp is None:
        return None
    elif 'BP' in paper_inform:
        bp = paper_inform['BP'][0]
        if bp == cr_bp:
            return True  # BP matched.
        else:
            return False  # Wrong BP.
    else:
        return None


def cr_match(paper_inform, cr):
    result = (doi_match(paper_inform, cr),
              au_match(paper_inform, cr),
              vl_match(paper_inform, cr),
              bp_match(paper_inform, cr))
    if False in result:
        return False
    elif result[0] is None:
        if result[1:] == (True, True, True):
            return True
        else:
            return False
    elif result[1] is True:
        return True
    else:
        return False


root_path = '' # type work path
sumup_path = os.path.join(root_path, 'SumUp')
sumupIDD_path = os.path.join(root_path, 'SumUp_indexed')
cr_ut_path = os.path.join(root_path, 'cr_database_ut')
cr_ut_refine_path = os.path.join(root_path, 'cr_database_ut_refined')
cr_path = os.path.join(root_path, 'cr_database')
cr_refine_path = os.path.join(root_path, 'cr_database_refined')
cr_index_path = os.path.join(root_path, 'cr_database_indexed')
org_path = os.path.join(root_path, 'org_database_indexed')
cr_match_path = os.path.join(root_path, 'cr_database_matched')
cnet_path = os.path.join(root_path, 'CNet')

Jlist = pd.read_excel(os.path.join(root_path,
                                   'ESIMasterJournalList-062018.xlsx'))
journal_FN_sub = dict(zip(Jlist['Full title'], Jlist['Category name']))
journal_short = dict(zip(Jlist['Title20'], list(zip(Jlist['Category name'], Jlist['Full title']))))
journal_short_bak = dict(zip(Jlist['Title29'], list(zip(Jlist['Category name'], Jlist['Full title']))))
journal_short.update(journal_short_bak)
journals = os.listdir(sumup_path)

# 001：Process the data under the SumUp folder, further decompose the data by volume number, and generate SumUp_ Indexed folder
print('No.1 Start!')
for journal in journals:
    sub = journal_FN_sub[journal]
    jn_path = os.path.join(sumup_path, journal)
    jn_index = {}
    out_count = {}
    for paper_inform in path_reader(jn_path):
        if 'PY' in paper_inform:
            py = paper_inform['PY'][0]
        else:
            py = paper_inform['EY'][0]
        if 'VL' in paper_inform:
            vl = 'V' + paper_inform['VL'][0]
        else:
            vl = 'Unknown'
        if py in jn_index:
            if vl in jn_index[py]:
                jn_index[py][vl].append(paper_inform)
                if len(jn_index[py][vl]) == 500:
                    name = '{}-{}.txt'.format(1+out_count[py][vl]*500, (1+out_count[py][vl])*500)
                    out_path = os.path.join(sumupIDD_path, sub, journal, py+'; '+vl)
                    buffer_writer(jn_index[py][vl], out_path, name)
                    jn_index[py][vl] = []
                    out_count[py][vl] += 1
            else:
                jn_index[py][vl] = [paper_inform]
                out_count[py][vl] = 0
        else:
            jn_index[py] = {}
            out_count[py] = {}
            jn_index[py][vl] = [paper_inform]
            out_count[py][vl] = 0
    for py in jn_index:
        for vl in jn_index[py]:
            if jn_index[py][vl] == []:
                continue
            name = '{}-{}.txt'.format(1+out_count[py][vl]*500, len(jn_index[py][vl])+out_count[py][vl]*500)
            out_path = os.path.join(sumupIDD_path, sub, journal, py+'; '+vl)
            buffer_writer(jn_index[py][vl], out_path, name)
            jn_index[py][vl] = []
print('No.1 End!')

# 002：Extract the "UT" and "CR" fields from the "SumUp_indexed" folder, decompose to the journals corresponding to "CR", and output to the "cr_database_ut" folder.
print('No.2 Start!')
cr_jn = {}
cr_jn_index = {}
N = 0
n = 0

for paper_inform in path_reader_2(sumupIDD_path):
    if 'CR' not in paper_inform:
        continue
    ut = paper_inform['UT'][0][4:]
    for cr in paper_inform['CR']:
        N += 1
        cr = ut + ': ' + cr
        cr = cr.upper()
        cr_inform = cr.split(', ')
        if len(cr_inform) < 3:
            continue
        if cr_inform[2] in journal_short:
            n += 1
            jn_ft = journal_short[cr_inform[2]][1].strip()
            if jn_ft in cr_jn:
                cr_jn[jn_ft].append(cr)
                if len(cr_jn[jn_ft]) == 50000:
                    cr_jn[jn_ft] = list(set(cr_jn[jn_ft]))
                if len(cr_jn[jn_ft]) == 50000:
                    if jn_ft in cr_jn_index:
                        batch_writer(cr_jn_index[jn_ft][0], cr_jn_index[jn_ft][1], 50000, '\n'.join(cr_jn[jn_ft]))
                        cr_jn[jn_ft] = []
                        cr_jn_index[jn_ft][1] += 1
                    else:
                        cr_jn_path = os.path.join(cr_ut_path, jn_ft)
                        if not os.path.exists(cr_jn_path):
                            os.makedirs(cr_jn_path)
                        cr_jn_index[jn_ft] = [cr_jn_path, 1]
                        batch_writer(cr_jn_path, 0, 50000, '\n'.join(cr_jn[jn_ft]))
                        cr_jn[jn_ft] = []
            else:
                cr_jn[jn_ft] = [cr]

for jn_ft in cr_jn:
    if cr_jn[jn_ft] == []:
        continue
    cr_jn[jn_ft] = list(set(cr_jn[jn_ft]))
    if jn_ft in cr_jn_index:
        batch_writer(cr_jn_index[jn_ft][0], cr_jn_index[jn_ft][1], 50000, '\n'.join(cr_jn[jn_ft]))
        cr_jn[jn_ft] = []
        cr_jn_index[jn_ft][1] += 1
    else:
        cr_jn_path = os.path.join(cr_ut_path, jn_ft)
        if not os.path.exists(cr_jn_path):
            os.makedirs(cr_jn_path)
        cr_jn_index[jn_ft] = [cr_jn_path, 1]
        batch_writer(cr_jn_path, 0, len(cr_jn[jn_ft]), '\n'.join(cr_jn[jn_ft]))
        cr_jn[jn_ft] = []
print('No.2 End!')

# 003：Extract data from "cr_database_ut", refine using "cr" as the primary index, and retain only local journal results, outputting to "cr_database_ut_refined".
print('No.3 Start!')
N = 0
for journal in journals:
    jn_path = os.path.join(cr_ut_path, journal)
    crs = []
    crsd = {}
    for file in file_getter(jn_path):
        f = open(file, encoding='utf-8')
        temp = f.read().splitlines()
        f.close()
        for item in temp:
            item = item.split(': ')
            if item[1] in crsd:
                crsd[item[1]].append(item[0])
            else:
                crsd[item[1]] = [item[0]]
    for cr in crsd:
        crs.append(', '.join(crsd[cr]) + ': ' + cr)
    N += len(crs)
    content = '\n'.join(crs)
    batch_writer_lite(cr_ut_refine_path, journal, content)
print('No.3 End!')

# 004：Extract only the CR field from "SumUp_indexed", remove duplicates, then break down to the corresponding journals, and output to the "cr_database" folder.
print('No.4 Start!')
cr_jn = {}
cr_jn_index = {}
N = 0
n = 0

for paper_inform in path_reader_2(sumupIDD_path):
    if 'CR' not in paper_inform:
        continue
    for cr in paper_inform['CR']:
        N += 1
        cr = cr.upper()
        cr_inform = cr.split(', ')
        if len(cr_inform) < 3:
            continue
        if cr_inform[2] in journal_short:
            n += 1
            jn_ft = journal_short[cr_inform[2]][1].strip()
            if jn_ft in cr_jn:
                cr_jn[jn_ft].append(cr)
                if len(cr_jn[jn_ft]) == 50000:
                    cr_jn[jn_ft] = list(set(cr_jn[jn_ft]))
                if len(cr_jn[jn_ft]) == 50000:
                    if jn_ft in cr_jn_index:
                        batch_writer(cr_jn_index[jn_ft][0], cr_jn_index[jn_ft][1], 50000, '\n'.join(cr_jn[jn_ft]))
                        cr_jn[jn_ft] = []
                        cr_jn_index[jn_ft][1] += 1
                    else:
                        cr_jn_path = os.path.join(cr_path, jn_ft)
                        if not os.path.exists(cr_jn_path):
                            os.makedirs(cr_jn_path)
                        cr_jn_index[jn_ft] = [cr_jn_path, 1]
                        batch_writer(cr_jn_path, 0, 50000, '\n'.join(cr_jn[jn_ft]))
                        cr_jn[jn_ft] = []
            else:
                cr_jn[jn_ft] = [cr]

for jn_ft in cr_jn:
    if cr_jn[jn_ft] == []:
        continue
    cr_jn[jn_ft] = list(set(cr_jn[jn_ft]))
    if jn_ft in cr_jn_index:
        batch_writer(cr_jn_index[jn_ft][0], cr_jn_index[jn_ft][1], 50000, '\n'.join(cr_jn[jn_ft]))
        cr_jn[jn_ft] = []
        cr_jn_index[jn_ft][1] += 1
    else:
        cr_jn_path = os.path.join(cr_path, jn_ft)
        if not os.path.exists(cr_jn_path):
            os.makedirs(cr_jn_path)
        cr_jn_index[jn_ft] = [cr_jn_path, 1]
        batch_writer(cr_jn_path, 0, len(cr_jn[jn_ft]), '\n'.join(cr_jn[jn_ft]))
        cr_jn[jn_ft] = []
print('No.4 End!')

# 005：Extract data from "cr_database", retaining only the local journal results, and output to the "cr_database_refined" folder.
print('No.5 Start!')
for journal in journals:
    jn_path = os.path.join(cr_path, journal)
    crs = []
    for file in file_getter(jn_path):
        f = open(file, encoding='utf-8')
        crs += f.read().splitlines()
        f.close()
        crs = list(set(crs))
    N += len(crs)
    content = '\n'.join(crs)
    batch_writer_lite(cr_refine_path, journal, content)
print('No.5 End!')

# 006：Extract data from "cr_database_refined", further decompose based on the year information in the CR field, and output to the "cr_database_indexed" folder.
print('No.6 Start!')
for file in os.listdir(cr_refine_path):
    f_path = os.path.join(cr_refine_path, file)
    jn_path = os.path.join(cr_index_path, file.split('.')[0])
    if not os.path.exists(jn_path):
        os.makedirs(jn_path)
    cr_index = {'others': []}
    f = open(f_path, encoding='utf-8')
    for cr in f.read().splitlines():
        year = cr.split(', ')[1]
        if year.isdigit():
            if year in cr_index:
                cr_index[year].append(cr)
            else:
                cr_index[year] = [cr]
        else:
            cr_index['others'].append(cr)
    f.close()
    for key in cr_index:
        if cr_index[key] == []:
            continue
        if key == 'others':
            print(file)
            print('Incorrect year information in CR field')
        content = '\n'.join(cr_index[key])
        batch_writer_lite(jn_path, key, content)
print('No.6 End!')

# 007：Extract data from "SumUp_indexed", organize by grouping according to time, and output to the "org_database_indexed" folder.
print('No.7 Start!')
N = 0
n = 0
for sub in os.listdir(sumupIDD_path):
    sub_path = os.path.join(sumupIDD_path, sub)
    for journal in os.listdir(sub_path):
        jn_path = os.path.join(sub_path, journal)
        pp_index = {}
        pp_count = {}
        out_count = {}
        for file in file_getter(jn_path):
            for paper in txt_spliter(file):
                paper_inform = paper_extractor(paper)
                tc = int(paper_inform['TC'][0])
                if tc == 0:
                    n += 1
                    continue
                if 'PY' not in paper_inform:
                    year = paper_inform['EY'][0]
                else:
                    year = paper_inform['PY'][0]
                N += 1
                content = ''.join(paper) + 'ER\n\n'
                if year in pp_index:
                    pp_index[year] += content
                    pp_count[year] += 1
                    if pp_count[year] == 500:
                        name = '{}-{}'.format(1+out_count[year]*500, (1+out_count[year])*500)
                        batch_writer_lite(os.path.join(org_path, journal, year), name, pp_index[year]+'EF\n')
                        pp_index[year] = "FN Enhanced WoS' Data\nVR 1.1\n"
                        pp_count[year] = 0
                        out_count[year] += 1
                else:
                    pp_index[year] = "FN Enhanced WoS' Data\nVR 1.1\n" + content
                    pp_count[year] = 1
                    out_count[year] = 0
        for year in pp_index:
            if pp_count[year] == 0:
                continue
            name = '{}-{}'.format(1+out_count[year]*500, pp_count[year]+out_count[year]*500)
            batch_writer_lite(os.path.join(org_path, journal, year), name, pp_index[year]+'EF\n')
print('No.7 End!')

# 008：
Compare the data between "cr_database_indexed" and "org_database_indexed", map the CR to the original document (UT), and output to the "cr_database_matched" folder.
print('No.8 Start!')
if not os.path.exists(cr_match_path):
    os.makedirs(cr_match_path)
N = 0
n = 0
for journal in os.listdir(org_path):
    jn_path = os.path.join(org_path, journal)
    jn_cr_path = os.path.join(cr_index_path, journal)
    cr_files = os.listdir(jn_cr_path)
    cr_matched = {}
    for year in os.listdir(jn_path):
        cr_file = year + '.txt'
        if cr_file not in cr_files:
            continue
        f = open(os.path.join(jn_cr_path, cr_file), encoding='utf-8')
        crs = f.read().splitlines()
        f.close()
        for paper_inform in path_reader(os.path.join(jn_path, year)):
            ut = paper_inform['UT'][0][4:]
            for cr in crs:
                if cr_match(paper_inform, cr):
                    N += 1
                    if cr in cr_matched:
                        n += 1
                        cr_matched[cr].append(ut)
                    else:
                        cr_matched[cr] = [ut]
    name = os.path.join(cr_match_path, journal+'.txt')
    f = open(name, mode='w', encoding='utf-8')
    for cr in cr_matched:
        line = cr + '; ' + ', '.join(cr_matched[cr]) + '\n'
        f.write(line)
    f.close()
print('No.8 End!')

# 009：Build citation networks.
# Add fields CS (in CR format), LC (local citation document accession number), NC (local citation count), RR (local reference count), and LR (local reference document accession number).
# Note: Suppose a document A corresponds to the same CR with documents B, C,... (due to the limitations of the cr_match function or imperfect data), it will make the NC in its corresponding Cnet data equal to 0 (regardless of what the TC is).
# Note: If a document's reference list has a CR that matches multiple documents, then delete that CR entry.
print('No.9 Start!')
buffer_size = 500
jn_crs = {}
for journal in journals:
    jn_crs[journal] = {}
    match_file = os.path.join(cr_match_path, journal+'.txt')
    f = open(match_file, encoding='utf-8')
    # CR-UT pair form：BORHO W, 1981, ABH MATH SEM HAMBURG, V51, P1, DOI 10.1007/BF02941207; A1981MM81700001
    matches = f.read().splitlines()
    for match in matches:
        match = match.split('; ')
        if len(match[1].split(', ')) > 1:
            print('Error in cr_database_matched file! in', journal, match[0])
            continue
        jn_crs[journal][match[0]] = match[1]

for journal in journals:
    sub = journal_FN_sub[journal]
    jn_path = os.path.join(sumup_path, journal)
    out_path = os.path.join(cnet_path, journal)
    cp_file = os.path.join(cr_ut_refine_path, journal+'.txt')
    match_file = os.path.join(cr_match_path, journal+'.txt')
    buffer = []
    buffer_index = 0
    
    f = open(cp_file, encoding='utf-8')
    cps = f.read().splitlines()
    f.close()
    f = open(match_file, encoding='utf-8')
    matches = f.read().splitlines()
    f.close()
    cr_cp = {}
    ut_cs = {}
    for cp in cps:
        cp = cp.split(': ')
        cr_cp[cp[1]] = cp[0].split(', ')
    for match in matches:
        match = match.split('; ')
        if match[1] in ut_cs:
            ut_cs[match[1]].append(match[0])
        else:
            ut_cs[match[1]] = [match[0]]
    for paper_inform in path_reader(jn_path): 
        paper_inform['JN'] = [journal] 
        paper_inform['SJ'] = [sub] 
        if 'AU' in paper_inform: 
            paper_inform['TS'] = [str(len(paper_inform['AU']))]
        else:
            paper_inform['TS'] = ['0']
        if 'PY' not in paper_inform:
            paper_inform['PY'] = paper_inform['EY']
        ut = paper_inform['UT'][0][4:] 
        tc = int(paper_inform['TC'][0]) 
        if tc:
            if ut in ut_cs:
                css = ut_cs[ut]
                paper_inform['CS'] = css
                paper_inform['LC'] = []
                for cs in css:
                    paper_inform['LC'] += cr_cp[cs]
                paper_inform['NC'] = [str(len(paper_inform['LC']))]
            else:
                paper_inform['NC'] = ['0']
        else:
            paper_inform['NC'] = ['0']
        if 'CR' in paper_inform:
            lr = []
            for cr in paper_inform['CR']:
                cr = cr.upper()
                cr_inform = cr.split(', ')
                if len(cr_inform) < 3:
                    continue
                if cr_inform[2] in journal_short:
                    jn_ft = journal_short[cr_inform[2]][1].strip()
                    if jn_ft not in journals:
                        continue
                else:
                    continue
                if cr in jn_crs[jn_ft]:
                    lr.append(jn_crs[jn_ft][cr])
            paper_inform['RR'] = [str(len(lr))]
            if lr:
                paper_inform['LR'] = lr
        else:
            paper_inform['RR'] = ['0']
        buffer.append(paper_inform)
        if len(buffer) == buffer_size:
            file_name = '{}-{}.txt'.format(1+buffer_index*buffer_size, (1+buffer_index)*buffer_size)
            buffer_writer(buffer, out_path, file_name)
            buffer = []
            buffer_index += 1
    if buffer:
        file_name = '{}-{}.txt'.format(1+buffer_index*buffer_size, len(buffer)+buffer_index*buffer_size)
        buffer_writer(buffer, out_path, file_name)
print('No.9 End!')
